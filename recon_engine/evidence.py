"""Auditor evidence pack builder.

Generates the multi-sheet Excel workbook that an external auditor would
request during a SOX walkthrough. The workbook is reproducible from the
contents of `audit.recon_runs` + `audit.recon_check_results` + the
relevant `glrecon_marts.*` tables, so it serves as objective evidence
that the daily control was operating effectively on a given date.

The file is built in-memory (`io.BytesIO`) so the Streamlit page can
serve it via ``st.download_button`` without writing to disk.

Sheets:
    1. Run Summary       -- one row, the audit.recon_runs record
    2. Check Results     -- one row per check (audit.recon_check_results)
    3. Control Account   -- top-N account-level breaks
    4. Transaction Level -- top-N matching engine breaks
    5. Manual JE Flags   -- SOX red flags hitting control accounts
    6. Sign-Off          -- blank template the controller signs and dates
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import Engine, text

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)

STATUS_FILLS = {
    "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "WARN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autofit(ws, max_width: int = 50) -> None:
    """Best-effort column auto-fit (openpyxl has no native autofit)."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            text_len = len(str(cell.value))
            if text_len > max_len:
                max_len = text_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ---------------------------------------------------------------------------
# Per-sheet builders
# ---------------------------------------------------------------------------

def _build_run_summary_sheet(ws, run: dict[str, Any]) -> None:
    ws.title = "Run Summary"
    ws["A1"] = "GL Reconciliation — Auditor Evidence Pack"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated at {datetime.now(UTC).isoformat(timespec='seconds')}"

    rows = [
        ("Run ID",              str(run["run_id"])),
        ("Business date",       str(run["business_date"])),
        ("Status",              run["status"]),
        ("Triggered by",        run["triggered_by"]),
        ("Started at (UTC)",    str(run["started_at"])),
        ("Finished at (UTC)",   str(run["finished_at"])),
        ("Git commit SHA",      run.get("git_commit_sha") or "—"),
        ("dbt manifest hash",   run.get("dbt_manifest_hash") or "—"),
        ("Evidence URL",        run.get("evidence_url") or "—"),
    ]
    start_row = 4
    for i, (label, value) in enumerate(rows):
        ws.cell(row=start_row + i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=start_row + i, column=2, value=value)

    # Source row counts breakdown.
    counts = run.get("source_row_counts") or {}
    if counts:
        ws.cell(row=start_row + len(rows) + 2, column=1, value="Source row counts").font = LABEL_FONT
        for i, (table, n) in enumerate(sorted(counts.items())):
            ws.cell(row=start_row + len(rows) + 3 + i, column=1, value=table)
            ws.cell(row=start_row + len(rows) + 3 + i, column=2, value=int(n))

    _autofit(ws)


def _build_check_results_sheet(ws, checks: list[dict[str, Any]]) -> None:
    ws.title = "Check Results"
    headers = ["Check Name", "Status", "Breaks Count", "Breaks Value (USD)", "Materiality (USD)"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for c in checks:
        row_idx = ws.max_row + 1
        ws.append([
            c["check_name"],
            c["status"],
            int(c["breaks_count"] or 0),
            float(c["breaks_value_usd"] or 0),
            float(c["materiality_usd"]) if c.get("materiality_usd") else None,
        ])
        # Color-code the Status cell.
        fill = STATUS_FILLS.get(c["status"])
        if fill:
            ws.cell(row=row_idx, column=2).fill = fill

    # Number formats.
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5).number_format = '"$"#,##0.00'
    _autofit(ws)


def _build_marts_sheet(
    ws,
    title: str,
    headers: list[str],
    rows: list[tuple],
    status_col: int | None = None,
) -> None:
    ws.title = title
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for row in rows:
        row_idx = ws.max_row + 1
        ws.append(list(row))
        if status_col is not None:
            status_value = ws.cell(row=row_idx, column=status_col).value
            fill = STATUS_FILLS.get(status_value)
            if fill:
                ws.cell(row=row_idx, column=status_col).fill = fill
    _autofit(ws)


def _build_signoff_sheet(ws, run_id: uuid.UUID, business_date: date) -> None:
    ws.title = "Sign-Off"
    ws["A1"] = "Auditor Sign-Off"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = f"Reconciliation run for business date {business_date} (run_id {run_id})."
    ws["A4"] = "I have reviewed the attached evidence and confirm that:"
    ws["A5"] = "  • The control was performed on the stated business date."
    ws["A6"] = "  • All material breaks have been investigated and either resolved or accepted."
    ws["A7"] = "  • The accept/reject decisions are documented in the issue tracker."

    rows = [
        ("Reviewed by",   ""),
        ("Title",         ""),
        ("Date",          ""),
        ("Signature",     ""),
        ("",              ""),
        ("Approved by",   ""),
        ("Title",         ""),
        ("Date",          ""),
        ("Signature",     ""),
    ]
    for i, (label, value) in enumerate(rows):
        ws.cell(row=10 + i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=10 + i, column=2, value=value)
        if label and not label.startswith("Signature"):
            ws.cell(row=10 + i, column=2).alignment = Alignment(horizontal="left")
    _autofit(ws)


# ---------------------------------------------------------------------------
# Public builder
# ---------------------------------------------------------------------------

def build_evidence_pack(
    engine: Engine,
    run_id: uuid.UUID,
    top_n_breaks: int = 100,
    marts_schema: str = "glrecon_marts",
) -> bytes:
    """Return the Excel workbook as raw bytes ready for download."""
    log.info("evidence.build.started", run_id=str(run_id))

    with engine.connect() as conn:
        run_row = conn.execute(
            text("SELECT * FROM audit.recon_runs WHERE run_id = :r"),
            {"r": run_id},
        ).mappings().one()

        check_rows = conn.execute(
            text("""
                SELECT check_name, status, breaks_count, breaks_value_usd, materiality_usd
                FROM audit.recon_check_results
                WHERE run_id = :r
                ORDER BY check_name
            """),
            {"r": run_id},
        ).mappings().all()

        ctrl_breaks = conn.execute(
            text(f"""
                SELECT entity_id, account_code, posting_date,
                       sl_balance_usd, gl_balance_usd, variance_usd, status
                FROM {marts_schema}.recon_control_account
                WHERE status <> 'PASS'
                ORDER BY abs(variance_usd) DESC
                LIMIT :n
            """),
            {"n": top_n_breaks},
        ).all()

        txn_breaks = conn.execute(
            text(f"""
                SELECT source_system, source_doc_id, account_code, entity_id,
                       sl_amount_usd, gl_amount_usd, amount_delta_usd, break_class, status
                FROM {marts_schema}.recon_transaction_level
                WHERE status = 'BREAK'
                ORDER BY abs(amount_delta_usd) DESC
                LIMIT :n
            """),
            {"n": top_n_breaks},
        ).all()

        manual_jes = conn.execute(
            text(f"""
                SELECT journal_id, journal_line_id, entity_id, business_date,
                       account_code, amount_usd, created_by, description
                FROM {marts_schema}.recon_manual_je_flag
                ORDER BY abs(amount_usd) DESC
                LIMIT :n
            """),
            {"n": top_n_breaks},
        ).all()

    wb = Workbook()
    # Sheet 1
    _build_run_summary_sheet(wb.active, dict(run_row))
    # Sheet 2
    _build_check_results_sheet(wb.create_sheet(), [dict(r) for r in check_rows])
    # Sheet 3
    _build_marts_sheet(
        wb.create_sheet(),
        title="Control Account",
        headers=["Entity", "Account", "Posting Date",
                 "SL Balance (USD)", "GL Balance (USD)", "Variance (USD)", "Status"],
        rows=[(r.entity_id, r.account_code, r.posting_date,
               float(r.sl_balance_usd or 0), float(r.gl_balance_usd or 0),
               float(r.variance_usd or 0), r.status)
              for r in ctrl_breaks],
        status_col=7,
    )
    # Sheet 4
    _build_marts_sheet(
        wb.create_sheet(),
        title="Transaction Level",
        headers=["Source", "Source Doc ID", "Account", "Entity",
                 "SL Amount (USD)", "GL Amount (USD)", "Delta (USD)",
                 "Break Class", "Status"],
        rows=[(r.source_system, r.source_doc_id, r.account_code, r.entity_id,
               float(r.sl_amount_usd or 0) if r.sl_amount_usd is not None else None,
               float(r.gl_amount_usd or 0) if r.gl_amount_usd is not None else None,
               float(r.amount_delta_usd or 0),
               r.break_class, r.status)
              for r in txn_breaks],
        status_col=9,
    )
    # Sheet 5
    _build_marts_sheet(
        wb.create_sheet(),
        title="Manual JE Flags",
        headers=["Journal ID", "Line", "Entity", "Business Date",
                 "Account", "Amount (USD)", "Created By", "Description"],
        rows=[(r.journal_id, r.journal_line_id, r.entity_id, r.business_date,
               r.account_code, float(r.amount_usd or 0),
               r.created_by, r.description)
              for r in manual_jes],
    )
    # Sheet 6
    _build_signoff_sheet(
        wb.create_sheet(),
        run_id=run_id,
        business_date=run_row["business_date"],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log.info(
        "evidence.build.complete",
        run_id=str(run_id),
        bytes=len(buf.getvalue()),
        sheets=len(wb.sheetnames),
    )
    return buf.getvalue()


def evidence_filename(business_date: date, run_id: uuid.UUID) -> str:
    return f"GLRecon_Evidence_{business_date.isoformat()}_{str(run_id)[:8]}.xlsx"


# Suppress unused-import warning if Decimal isn't referenced elsewhere yet.
_ = Decimal
