"""Tests for ``recon_engine.evidence`` — the Excel auditor evidence pack."""

from __future__ import annotations

import io
import os
import uuid
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from recon_engine.audit import AuditTrailWriter, CheckResult
from recon_engine.evidence import build_evidence_pack, evidence_filename

# ---------------------------------------------------------------------------
# Pure unit tests
# ---------------------------------------------------------------------------

def test_evidence_filename_format():
    rid = uuid.UUID("11111111-2222-3333-4444-555555555555")
    name = evidence_filename(date(2026, 5, 3), rid)
    assert name == "GLRecon_Evidence_2026-05-03_11111111.xlsx"


# ---------------------------------------------------------------------------
# Integration tests (live Postgres)
# ---------------------------------------------------------------------------

def _live_engine():
    user = os.environ.get("POSTGRES_USER", "glrecon")
    password = os.environ.get("POSTGRES_PASSWORD", "glrecon")
    host = os.environ.get("POSTGRES_HOST", "localhost")
    port = os.environ.get("POSTGRES_PORT", "5432")
    db = os.environ.get("POSTGRES_DB", "glrecon")
    dsn = f"postgresql+psycopg://{user}:{password}@{host}:{port}/{db}"
    eng = create_engine(dsn, future=True)
    try:
        with eng.connect() as conn:
            conn.execute(text("SELECT 1"))
    except Exception as exc:
        pytest.skip(f"Postgres not reachable ({exc.__class__.__name__})")
    return eng


@pytest.fixture
def engine():
    return _live_engine()


@pytest.fixture
def seeded_run(engine):
    """Create a run + per-check rows that the evidence builder can read."""
    writer = AuditTrailWriter(engine)
    run_id = writer.start_run(
        business_date=date(2026, 5, 3),
        triggered_by="manual:test_evidence",
    )
    writer.finalize(
        run_id=run_id,
        checks=[
            CheckResult("control_account", "FAIL",
                        breaks_count=12, breaks_value_usd=Decimal("999.99")),
            CheckResult("transaction_level", "WARN",
                        breaks_count=3, breaks_value_usd=Decimal("12.34")),
            CheckResult("roll_forward", "PASS"),
        ],
        source_row_counts={"raw.gl_journal": 100050, "raw.ap_invoices": 10000},
        dbt_manifest_hash="testhash",
    )
    return run_id


def test_evidence_pack_contains_all_six_sheets(engine, seeded_run):
    data = build_evidence_pack(engine, run_id=seeded_run, top_n_breaks=5)
    assert isinstance(data, bytes)
    assert len(data) > 1024, "evidence pack should be non-trivial in size"

    wb = load_workbook(io.BytesIO(data))
    expected = {
        "Run Summary",
        "Check Results",
        "Control Account",
        "Transaction Level",
        "Manual JE Flags",
        "Sign-Off",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_run_summary_sheet_includes_provenance(engine, seeded_run):
    data = build_evidence_pack(engine, run_id=seeded_run, top_n_breaks=5)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Run Summary"]
    cells = [str(c.value) for row in ws.iter_rows(values_only=False) for c in row if c.value is not None]
    text_blob = " ".join(cells)
    assert str(seeded_run) in text_blob
    assert "testhash" in text_blob
    assert "raw.gl_journal" in text_blob
    assert "100050" in text_blob


def test_check_results_sheet_has_all_three_checks(engine, seeded_run):
    data = build_evidence_pack(engine, run_id=seeded_run, top_n_breaks=5)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Check Results"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    check_names = {row[0] for row in rows if row[0]}
    assert check_names == {"control_account", "transaction_level", "roll_forward"}


def test_signoff_sheet_has_blank_template(engine, seeded_run):
    data = build_evidence_pack(engine, run_id=seeded_run)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Sign-Off"]
    labels = [str(row[0].value) for row in ws.iter_rows() if row[0].value]
    assert any("Reviewed by" in lbl for lbl in labels)
    assert any("Approved by" in lbl for lbl in labels)
    assert any("Signature" in lbl for lbl in labels)
